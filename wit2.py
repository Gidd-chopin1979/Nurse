#30 file 8:26
#rowdataをmacroの方に移して，区切り番号を振っていく
import time
t_start = time.time()

import datetime
import openpyxl as px
from mymodule import my_round #負の値の丸め込みに使用
from mymodule import hhmmss #hhmmss>s
from decimal import Decimal, ROUND_HALF_UP, ROUND_HALF_EVEN #floatの四捨五入

xxx_name = ['001','002','003','004','005','006','007','008','009','010','011','012','013','014','015','016','017','018','019','020','021','022','023','024','025','026','027','028','029','030']
Except_List = [] #listの要素数が4つ揃っていないxxx #007,14,16,17,19,20,24,26,29
ErrRow_List = [] #上のse_RowList

for i in range(0,30): #range(0,30)
    witxl_path = './Result/001-030/1st_Wit_' + xxx_name[i] + '.xlsx'
    fxl_path = './Result/xlsx_template/failed/pre_Wit_' + xxx_name[i] + '.xlsx'

    wb = px.load_workbook(witxl_path)
    wbf = px.load_workbook(fxl_path)

    wsM = wb['macro']
    wsP = wb['paste-rowdata']
    ws12 = wb['paste-1and2']
    max_row = wsP.max_row

    wsfM = wbf['macro']
    wsf12 = wbf['paste-1and2']
    max_rowf = wsfM.max_row

    #生データの一部(timeとx)を'macro'へコピー
    for k in range(3,max_row+1): #dataは3行目から
        wsM.cell(row=k, column=2).value = wsP.cell(row=k, column=2).value #生データB列をmacroのB列へ
        wsM.cell(row=k, column=3).value = my_round.main(wsP.cell(row=k, column=9).value) #生データI列を丸めてmacroのD列へ

    #'macro'中，該当するデータに区切り番号を付与
    se_name = ['1','2','3','4'] #start-endの名前
    se_RowList = [] #区切りを把握しておく

    fill_a = px.styles.PatternFill(patternType='solid', fgColor='FFDC00', bgColor='FFDC00') #赤っぽい色に塗りつぶし
    fill_b = px.styles.PatternFill(patternType='solid', fgColor='00DCFF', bgColor='00DCFF') #青っぽい色に塗りつぶし
    
    for l in range(0,len(se_name)): #1回目開始終了，2回目開始終了の計4つ
        for j in range(3,max_row+1):

            prtsc_time = str(wb[se_name[l]]['V3'].value)
            rawX = str(wsM.cell(row=j, column=2).value)
            raw_time = str(rawX)

            prtsc_angle = str(wb[se_name[l]]['V4'].value) #小数第２位
            raw_angle = str(wsM.cell(row=j, column=3).value) #小数第４位

            if prtsc_time in raw_time: #macroのtimeに，PrtScのtimeが部分的に入っていればそのセルを塗りつぶし
                wsM.cell(row=j, column=2).fill = fill_a
            if prtsc_angle in raw_angle: #角度についても同様
                wsM.cell(row=j, column=3).fill = fill_b
            if prtsc_time in raw_time and prtsc_angle in raw_angle:
                wsM.cell(row=j, column=1).value = se_name[l] #timeと角度が両方一致している行に区切り番号を割り振る
                se_RowList.append(j) #区切りに該当する行番号をlistに格納

    #ここから'paste-1and2': 'macro'において1-2, 3-4に該当するデータを'paste-1and2'に切り出す
    print(xxx_name[i], ':', se_RowList)
    #が，listの要素が4つ揃っていないものがある．(上のコードが完璧でないことに起因)
    #そのxxxを捉えて，揃えたfileからコピーする
    try: #4つ揃っているもの
        trial_1 = se_RowList[1] - se_RowList[0] + 1 #1回目のデータ数
        trial_2 = se_RowList[3] - se_RowList[2] + 1 #2回目のデータ数

        ws12.cell(row=6, column=2).value = trial_1 #データ数をセルに記入
        ws12.cell(row=6, column=9).value = trial_2

        for l in range(trial_1): #1回目
            ws12.cell(row=l+6, column=3).value = wsM.cell(row=l+se_RowList[0], column=2).value #隣のシートからのコピペ
            ws12.cell(row=l+6, column=4).value = wsM.cell(row=l+se_RowList[0], column=3).value

        for l in range(trial_2): #2回目
            ws12.cell(row=l+6, column=10).value = wsM.cell(row=l+se_RowList[2], column=2).value
            ws12.cell(row=l+6, column=11).value = wsM.cell(row=l+se_RowList[2], column=3).value

    except IndexError as e: #4つ揃っていないもの
        print('要素数が足りてない:', e)
        Except_List.append(i)
        ErrRow_List.append([se_RowList])

        se_RowList_F = [] #揃ってるfileから要素数を取ってくる
        for k in se_name:
            for l in range(3, max_rowf+1):
                rawfX = str(wsfM.cell(row=l, column=1).value) #'macro'1列目に"1,2,3,4"が記載されている
                if rawfX == k:
                    se_RowList_F.append(l)

        print(xxx_name[i], '(修正)', ':', se_RowList_F)
        trial_f1 = se_RowList_F[1] - se_RowList_F[0] + 1 #1回目のデータ数
        trial_f2 = se_RowList_F[3] - se_RowList_F[2] + 1 #2回目のデータ数

        #paste-1and2を丸ごと複製
        ws12.cell(row=6, column=2).value = trial_f1 #データ数をセルに記入
        ws12.cell(row=6, column=9).value = trial_f2

        for t1 in range(trial_f1): #1回目
            ws12.cell(row=t1+6, column=3).value = wsf12.cell(row=t1+6, column=3).value
            ws12.cell(row=t1+6, column=4).value = wsf12.cell(row=t1+6, column=4).value

        for t2 in range(trial_f2): #2回目
            ws12.cell(row=t2+6, column=10).value = wsf12.cell(row=t2+6, column=10).value
            ws12.cell(row=t2+6, column=11).value = wsf12.cell(row=t2+6, column=11).value
        
        wb.save(witxl_path) #Errfileのsave

        if xxx_name[i] != '030': #xxx=030なら終了
            continue
        else:
            break

    wb.save(witxl_path) #正常fileのsave

for el in range(len(Except_List)):
    print(xxx_name[Except_List[el]], 'のlistは', ErrRow_List[el])

t_end = time.time()
print(t_end-t_start, '要しました')
'''
001
[2713, 5629, 7974, 9688]
002
[9452, 12014, 14404, 15852]
003
[3436, 5705, 7533, 9225]
004
[2158, 5126, 7189, 8627]
005
[6676, 9652, 11891, 14916]
006
[6797, 8245, 11421, 14214]
007
[4836, 6787, 9054]
007 はlistの要素数が足りてない: list index out of range
008
[2446, 5076, 7094, 8779]
009
[2489, 4444, 6733, 7945]
010
[1988, 4134, 8738, 8740]
011
[3270, 4950, 7393, 9633]
012
[1902, 4190, 6249, 8173]
013
[4270, 7106, 9393, 11993]
014
[4960, 8109, 10914]
014 はlistの要素数が足りてない: list index out of range
015
[5015, 7543, 9791, 12453]
016
[4338, 6561, 10209]
016 はlistの要素数が足りてない: list index out of range
017
[8666, 13013]
017 はlistの要素数が足りてない: list index out of range
018
[5843, 7712, 9778, 11854]
019
[5671, 9329, 12315]
019 はlistの要素数が足りてない: list index out of range
020
[8047]
020 はlistの要素数が足りてない: list index out of range
021
[11678, 14579, 17007, 19373]
022
[4535, 5728, 7294, 9224]
023
[4504, 6247, 8259, 10862]
024
[6549, 10487, 12617]
024 はlistの要素数が足りてない: list index out of range
025
[3645, 6518, 8554, 11973]
026
[7293, 9766, 14196]
026 はlistの要素数が足りてない: list index out of range
027
[4660, 7676, 10378, 12408]
028
[4944, 8440, 10610, 13614]
029
[23338, 25757, 27632]
029 はlistの要素数が足りてない: list index out of range
030
[7924, 9699, 11757, 14492]
007 のlistは [[6549, 10487, 12617]]
'''