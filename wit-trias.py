#wit dataを0.1s区切りの綺麗な並びにしてcsvで出力
import time
t_start = time.time()

import os
import csv
import math
import openpyxl as px
from scipy import stats
from mymodule import hhmmss #hhmmss>s

xxx_name = ['001','002','003','004','005','006','007','008','009','010','011','012','013','014','015','016','017','018','019','020','021','022','023','024','025','026','027','028','029','030']

for i in range(0,30):
    for j in [1,2]: #1つのxxxにつき2つ(1回目2回目)
        print(xxx_name[i],'_',j)
        xint_list = [] #存在しない切り捨て整数
        xrow_list = [] #空欄の行

        witxl_path = './Result/001-030/1st_Wit_' + xxx_name[i] + '.xlsx'
        fxl_path = './Result/xlsx_template/failed/pre_Wit_' + xxx_name[i] + '.xlsx'

        tri_temp_path = './Result/xlsx_template/Wit-TRIAS.xlsx'
        tri_path = './Result/001-030/W-TRIAS/xlsx/Wit-TRIAS_' + xxx_name[i] + '_' + str(j) + 'st.xlsx'
        csv_path = os.path.join('./Result/001-030/W-TRIAS/csv/Wit-TRIAS_' + xxx_name[i] + '_' + str(j) + 'st.csv')#,で連結

        wbW = px.load_workbook(witxl_path)
        wbF = px.load_workbook(fxl_path)
        wbT = px.load_workbook(tri_temp_path)

        ws12 = wbW['paste-1and2']
        ws12F = wbF['paste-1and2']
        wsC = wbW['corr-factor']
        wsA = wbT['arrange']

        Li = [0,7]
        L = ws12.cell(row=6, column=Li[j-1]+2).value #data数

        #wit[paste-1and2]のangleと経過時間(DE6-)をBC6へコピー(経過時間は=(C6-$C$6)*24*60*60 だが掛け算はhhmmss.main()で実行される)
        for k in range(0,L): #L
            try:
                wt0 = hhmmss.main(ws12.cell(row=6,column=Li[j-1]+3).value) #$C$6
                wt = hhmmss.main(ws12.cell(row=k+6,column=Li[j-1]+3).value) #C6 - 
                wt_c = wt-wt0
                wsA.cell(row=k+6,column=2).value = ws12.cell(row=k+6,column=Li[j-1]+4).value #angle
                wsA.cell(row=k+6,column=3).value =  wt_c #経過時間
                wsA.cell(row=k+6,column=4).value = math.floor(wt_c) #DにC(元E)の切り捨て整数math.floorを記入
            except AttributeError as ae:#paste-1and2に値をコピーしてないwitfile
                print(ae)
                for kf in range(0,L):
                    wt0 = hhmmss.main(ws12F.cell(row=6,column=Li[j-1]+3).value) #$C$6
                    wt = hhmmss.main(ws12F.cell(row=kf+6,column=Li[j-1]+3).value) #C6 - 
                    wt_c = wt-wt0
                    wsA.cell(row=kf+6,column=2).value = ws12F.cell(row=kf+6,column=Li[j-1]+4).value #angle
                    wsA.cell(row=kf+6,column=3).value =  wt_c #経過時間
                    wsA.cell(row=kf+6,column=4).value = math.floor(wt_c) #DにC(元E)の切り捨て整数math.floorを記入

        #macro: 1秒あたりのデータが、10個以上なら以降無視，次へ．10以下なら直前の値をコピー(B>J)．
        #各整数の値がいくつあるのか調べる len(n_list)
        D_list = [wsA.cell(row=val+6,column=4).value for val in range(0,L)] #切り捨て整数のリスト[0,0,...,1,1,1...]
        d_list = [i for i in range(max(D_list)+1)] #各整数が1個ずつ入ったリスト[0,1,2,...]
        
        for l in d_list: 
            i_list = [l for s in range(D_list.count(l))] #D_listを整数ごとに分割 各整数の個数"D_list.count"分，各整数"l"を入れる
            l_ilis = len(i_list) #=D_list.count(i) その整数が何個あるか

            try:
                rb_ini = 6 + D_list.index(l) #6行目からスタート
                rj_ini = l*10 + 6 #j列は必ず10個ずつ
            except ValueError as ve: #""is not in listはlist.index("")のerror
                #print(ve)
                xint_list.append(l)

            if l_ilis >= 10: #その整数の個数が10個もしくはそれ以上ならそのままコピー(10で絶対とまるので以上でも問題ない)
                for r in range(10): #i=0の時，range(6,15+1)
                    wsA.cell(row=r+rj_ini,column=10).value = wsA.cell(row=r+rb_ini,column=2).value
            else: #10個より少ないなら, 直前の値をコピー
                for s in range(l_ilis): #7個の時，0123456
                    wsA.cell(row=s+rj_ini,column=10).value = wsA.cell(row=s+rb_ini,column=2).value
                for t in range(l_ilis, 10): #789
                    wsA.cell(row=t+rj_ini,column=10).value = wsA.cell(row=t+rj_ini-1,column=10).value #直前の値コピー
        
        #主要な値をxlsxに入れとく
        wsA.cell(row=4,column=4).value = d_list[-1] #整数の最大値
        for N in d_list:
            wsA.cell(row=N+6,column=5).value = N #各整数
            wsA.cell(row=N+6,column=6).value = D_list.count(N) #各整数の個数
        
        #wit[corr-factor]の傾き，切片(L6,7)を再計算
        if j == 1:
            sheet_names = {'A': ' (1)','B': ' (2)','C': ' (3)','D': ' (4)','E': ' (5)'}
        else:
            sheet_names = {'F': ' (6)','G': ' (7)','H': ' (8)','I': ' (9)','J': ' (10)'}
        x = [wbW[sheet_name]['V4'].value for sheet_name in sheet_names] #Wit_angle
        y = [wbW[sheet_name]['V5'].value for sheet_name in sheet_names] #kinovea_angle
        slope, intercept, r, p, std_err = stats.linregress(x,y)

        #I2,3に貼付け
        wsA['I2'].value = slope
        wsA['I3'].value = intercept
        sl = float(slope)
        icept = float(intercept)

        #補正値としてcol=9に記述
        for m in range(0,L):
            try:
                j_val = float(wsA.cell(row=m+6,column=10).value)
                wsA.cell(row=m+6,column=9).value = j_val * sl + icept #I6から=J6*$I$2+$I$3
            except TypeError as e:
                #print(m,'行目は，欠番か最後の値なので空欄です．')
                xrow_list.append(m)

        #HIを5から最後?までcsvで保存 シートを作ってそれをcsvとして保存する
        wbT.create_sheet(title='csv', index=1)
        ws_csv = wbT['csv']
        for n in range(0,L):
            ws_csv.cell(row=n+1,column=1).value = wsA.cell(row=n+5,column=8).value
            ws_csv.cell(row=n+1,column=2).value = wsA.cell(row=n+5,column=9).value

        m_row = range(1, ws_csv.max_row + 1)
        m_col = [1,2]

        with open(csv_path, 'w', encoding='Shift-JIS', newline="") as f:
            writer = csv.writer(f)
            writer.writerows([[ws_csv.cell(row=R, column=C).value for C in m_col] for R in m_row])
            #writer.writerow([str(acell.value or '') for acell in row]) #セルの値がNone, 空, 0, falseなら''
        f.close()
        
        #欠番の整数をコピー
        print(xxx_name[i],'_',j,'には',xint_list,'が存在しません．')
        wsA.cell(row=2,column=10).value = xint_list

        #空欄の行をコピー
        print(xxx_name[i],'_',j,'の','行目は，欠番か最後の値なので空欄です．')
        wsA.cell(row=3,column=10).value = xrow_list

        wbT.save(tri_path)

t_end = time.time()
print(t_end-t_start, '秒要しました')    