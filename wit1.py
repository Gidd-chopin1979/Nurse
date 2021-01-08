#kinoveaの画像含む，15img/1fileを貼り付ける
#witのelsx'paste-rowdata'にwitの生データ.txtを貼り付ける
#読み込み速度の観点から，読み込む列を限定しても良いかもしれない
import openpyxl as px
from openpyxl.drawing.image import Image
from mymodule import my_round

#xxx_name = '002' #ここを範囲指定するかは考え中．その場合は全体をfor
xxx_name = ['001','002','003','004','005','006','007','008','009','010','011','012','013','014','015','016','017','018','019','020','021','022','023','024','025','026','027','028','029','030']

for i in range(0,1):
    print(xxx_name[i])

    wtemp_path = './Result/xlsx_template/1st_Wit_xxx.xlsx'
    witxl_path = './Result/001-030/Wit/1st_Wit_' + xxx_name[i] + '.xlsx'
    imgfile_path = './Data/Image_001-030/img_' + xxx_name[i] + '/' + xxx_name[i] + '_'
    txt_path = './Data/Wit_001-030/' + xxx_name[i] + 'Data.txt'

    wb = px.load_workbook(wtemp_path)

    sheet_names1 = ['0','1','2','3','4']
    sheet_names2 = {'A': ' (1)','B': ' (2)','C': ' (3)','D': ' (4)','E': ' (5)','F': ' (6)','G': ' (7)','H': ' (8)','I': ' (9)','J': ' (10)'}

    wsM = wb['macro']
    wsP = wb['paste-rowdata']
    max_row = wsP.max_row

    #グラフタイトル名の参照用セルに書き込み
    wb['graph']['A1'].value = '1次_Wit_' + xxx_name[i]
    wb['graph']['L1'].value = '1次_Wit補正_' + xxx_name[i]

    #past_img.py
    for sheet_name in sheet_names1:

        ws = wb[sheet_name]
        img = Image(imgfile_path + sheet_name + '.JPG')
        img.width = 1378 #35 cm (100dpi環境下)
        img.height = 775 #19.69 cm

        ws.add_image(img, 'A1')

    for s_name, f_name in sheet_names2.items():

        ws = wb[s_name]
        img = Image(imgfile_path + f_name + '.png')
        img.width = 1378 #35 cm (100dpi環境下)
        img.height = 775 #19.69 cm

        ws.add_image(img, 'A1')

    #paste_tx.py
    with open(txt_path, encoding='UTF-8') as file: #一部文字化けしたので，encode指定
        lines = file.readlines()

    wsP.cell(1,1).value = lines[0] #txt file １行目は独立したものだったので．

    ttl_list = lines[1].split() #2行目は各列のタイトル．各要素に分割．
    for k in range(1,len(lines)): #このfileの行数分繰返し
        for j in range(1,len(ttl_list)+1): #何列のデータか，その分繰返し
            wsP.cell(row=k+1, column=j).value = lines[k].split()[j-1] #A2スタートで各リストの要素0から記述

    '''ここにalt_macro.pyを入れると実行されない．>なぜかわからん．load saveをもう１セット入れても無理だった．
    for k in range(3,max_row+1): #このfileの行数分繰返し
        wsM.cell(row=k, column=2).value = wsP.cell(row=k, column=2).value #生データB列をmacroのB列へ
        wsM.cell(row=k, column=3).value = my_round.main(wsP.cell(row=k, column=9).value) #生データI列を四捨五入してmacroのD列へ
    '''
    wb.save(witxl_path)