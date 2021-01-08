#witのelsx'paste-rowdata'にwitの生データ.txtを貼り付ける
#読み込み速度の観点から，読み込む列を限定しても良いかもしれない
import csv
import math
import openpyxl as px
from mymodule import hhmmss
from decimal import Decimal, ROUND_HALF_UP, ROUND_HALF_EVEN #floatの四捨五入

xxx_name = '002' #ここを範囲指定するかは考え中．その場合は全体をfor
bonxl_path = './Result/1st_Bonaly_' + xxx_name + '.xlsx'
witxl_path = './Result/1st_Wit_' + xxx_name + '.xlsx'
hr_path = './Data/Bonaly_001-030/1st_Bonaly_' + xxx_name + '/' + xxx_name +'.csv'
acl_path = './Data/Bonaly_001-030/1st_Bonaly_' + xxx_name + '/' + xxx_name +'_acl.csv'

#bonaly
wbB = px.load_workbook(bonxl_path)
wsH = wbB['hrate']
wsA = wbB['accel']
ws12 = wbB['1and2']
wsHG = wbB['hrate_graph']
maxRowH = wsH.max_row
maxRowA = wsA.max_row

#wit
wbW = px.load_workbook(witxl_path)
wsM = wbW['macro']
maxRowM = wsM.max_row

#.csvを'hrate'へ貼付け
with open(hr_path, encoding='UTF-8') as file: #txtを開くときと同じ
    line1d = csv.reader(file) #readlineでも可能だが，csv.readerを使う方が勝手がよさそう
    line2d = [row for row in line1d] #二次元配列＝リストのリストとして取得

for i in range(len(line2d)): #このfileの行数分繰返し
    L_data = line2d[i] 
    for j in range(len(L_data)): #各リストの要素数分だけ繰返し
        wsH.cell(row=i+1, column=j+1).value = line2d[i][j]

#_acl.csvを'accel'へ貼付け
with open(acl_path, encoding='UTF-8') as file: #txtを開くときと同じ
    line1D = csv.reader(file) 
    line2D = [row for row in line1D]

for i in range(len(line2D)):
    L_data = line2D[i] 
    for j in range(len(L_data)):
        wsA.cell(row=i+1, column=j+1).value = line2D[i][j]

#'accel'のA列のデータをhh:mm:ssに返還してG列へ
del line2D[0]
for i in range(len(line2D)):
    wsA.cell(row=i+2, column=7).value = float(line2D[i][0])/(24*3600)

#hrateの開始時間に該当するセルをaccelから探し出し，その経過時刻をhrateのEventに記入
fill_a = px.styles.PatternFill(patternType='solid', fgColor='FFDC00', bgColor='FFDC00')
start_list = []

for i in range(len(line2D)):
    wsH_val = str(hhmmss.main(wsH['A12'].value))
    wsA_val = str(wsA.cell(row=i+1, column=1).value)
    if wsH_val in wsA_val: 
        wsA.cell(row=i+1, column=7).fill = fill_a
        start_list.append(i+1) #list"line2D"は要素を1つ消していて，行数が合わなくなるので+1

wsA.cell(row=start_list[0], column=9).value = 0 #accelのI列に0を記入
wsH.cell(row=12, column=2).value = wsA.cell(row=start_list[0], column=8).value #accelのH列「経過時間」をhrateのB12に記入

#hrateの経過時刻を+2ずつして最下段までfilldown
for i in range(len(line2d)-12):
    wsH.cell(row=i+13, column=2).value = wsH.cell(row=i+12, column=2).value + 2 #hrateのB列を12行目から2ずつ足していく

#witの区切り番号をse_RowListに格納
se_name = ['1','2','3','4'] #start-endの名前
se_RowList = [] #区切りを把握しておく

for i in range(1,len(se_name)+1): #1回目開始終了，2回目開始終了の計4つ
    for j in range(1,maxRowM+1):
        if wsM.cell(row=j, column=1).value == i:
            se_RowList.append(j) #区切りに該当する行番号をlistに格納
print(se_RowList)

#witの区切りに一致する時刻をaccelから探して区切る
bw_RowList = [] #wit macroのデータと一致したbonaly accelのデータの行

for i in range(1,len(se_name)+1):
    wsM_val = float(hhmmss.main(wsM.cell(row=se_RowList[i-1], column=2).value))            
    for j in range(2,len(line2D)+1): #line2Dは要素を1つ消されているので+1
        wsA_val = Decimal(wsA.cell(row=j, column=1).value).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP) #roundはfloatを正確に取得できず，0.5の丸め方が一定方向でない
        if str(wsA_val) in str(wsM_val):
            wsA.cell(row=j, column=7).fill = fill_a 
            bw_RowList.append(j)
            wsA.cell(row=j, column=9).value = i
print(bw_RowList)

#1and2に1,2回目に該当する経過時間と角度のデータを移行
trial_1 = bw_RowList[1] - bw_RowList[0] + 1 #1回目のデータ数
trial_2 = bw_RowList[3] - bw_RowList[2] + 1 #2回目のデータ数
ws12.cell(row=1,column=4).value = str(trial_1)
ws12.cell(row=1,column=9).value = str(trial_2)

for i in range(trial_1):
    x = float(wsA.cell(row=i+bw_RowList[0], column=2).value)
    y = float(wsA.cell(row=i+bw_RowList[0], column=4).value)
    ws12.cell(row=i+3, column=3).value = math.atan2(y,x)*180/math.pi*(-1) #角度計算(式のコピーを防ぐ)
    ws12.cell(row=i+3, column=2).value = wsA.cell(i+bw_RowList[0], column=8).value #経過時間のコピー
    '''for j in range(maxRowH-12+1):
        ws12_valA = Decimal(ws12.cell(row=i+3, column=2).value).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP) #float
        wsH_valA = Decimal(wsH.cell(row=j+12, column=2).value).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP) #float
        if ws12_valA == wsH_valA:
            ws12.cell(row=i+3, column=4).value = wsH.cell(row=j+12, column=4).value'''

for j in range(maxRowH-12+1):
    for k in range(trial_1):
        ws12_valA = Decimal(ws12.cell(row=k+3, column=2).value).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP) #float
        wsH_valA = Decimal(wsH.cell(row=j+12, column=2).value).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP) #float
        if ws12_valA == wsH_valA:
            ws12.cell(row=k+3, column=4).value = wsH.cell(row=j+12, column=4).value

for i in range(trial_2):
    x = float(wsA.cell(row=i+bw_RowList[3], column=2).value)
    y = float(wsA.cell(row=i+bw_RowList[3], column=4).value)
    ws12.cell(row=i+3, column=8).value = math.atan2(y,x)*180/math.pi*(-1)
    ws12.cell(row=i+3, column=7).value = wsA.cell(i+bw_RowList[3], column=8).value #経過時間のコピー
    for j in range(maxRowH-12+1):
        ws12_valB = Decimal(ws12.cell(row=i+3, column=7).value).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP) #float
        wsH_valB = Decimal(wsH.cell(row=j+12, column=2).value).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP) #float
        if ws12_valB == wsH_valB:
            ws12.cell(row=i+3, column=9).value = wsH.cell(row=j+12, column=4).value

#1and2からhrate_graphへコピー
for i in range(trial_1):
    wsHG.cell(row=i+3,column=2).value = ws12.cell(row=i+3,column=4).value
    wsHG.cell(row=i+3,column=3).value = ws12.cell(row=i+3,column=5).value
for i in range(trial_2):
    wsHG.cell(row=i+3,column=5).value = ws12.cell(row=i+3,column=9).value
    wsHG.cell(row=i+3,column=6).value = ws12.cell(row=i+3,column=10).value

wbB.save(bonxl_path)