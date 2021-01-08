#witのelsx'paste-rowdata'にwitの生データ.txtを貼り付ける
#読み込み速度の観点から，読み込む列を限定しても良いかもしれない
import openpyxl as px

xxx_name = '002' #ここを範囲指定するかは考え中．その場合は全体をfor
xlfile_path = './Result/1st_Wit_' + xxx_name + '.xlsx'

wb = px.load_workbook(xlfile_path, data_only=True) #セルに入った式ではなくその結果の値を読むようにする

wsM = wb['macro']
wsP = wb['paste-rowdata']
max_row = wsP.max_row

wsSE = wb['start-end']
time_list = [5,6,8,9]

fill_a = px.styles.PatternFill(patternType='solid', fgColor='FFDC00', bgColor='FFDC00')
fill_b = px.styles.PatternFill(patternType='solid', fgColor='00DCFF', bgColor='00DCFF')

for i in range(0,len(time_list)): #1回目開始終了，2回目開始終了の計4つ
    for j in range(3,max_row+1):
        if str(wsSE.cell(row=time_list[i],column=5).value) in str(wsM.cell(row=j, column=2).value):
            wsM.cell(row=j, column=2).fill = fill_a
        if str(wsSE.cell(row=time_list[i],column=6).value) in str(wsM.cell(row=j, column=4).value):
            wsM.cell(row=j, column=4).fill = fill_b
        if str(wsSE.cell(row=time_list[i],column=5).value) in str(wsM.cell(row=j, column=2).value) and str(wsSE.cell(row=time_list[i],column=6).value) in str(wsM.cell(row=j, column=4).value):
            wsM.cell(row=j, column=3).value = i + 1

wb.save(xlfile_path)