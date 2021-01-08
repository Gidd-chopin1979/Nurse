#witのelsx'paste-rowdata'にwitの生データ.txtを貼り付ける
#読み込み速度の観点から，読み込む列を限定しても良いかもしれない
import openpyxl as px

xxx_name = '002' #ここを範囲指定するかは考え中．その場合は全体をfor
xlfile_path = './Result/1st_Wit_' + xxx_name + '.xlsx'
txt_path = './Data/Wit_001-030/' + xxx_name + 'Data.txt'

wb = px.load_workbook(xlfile_path)
ws = wb['paste-rowdata']

with open(txt_path, encoding='UTF-8') as file: #一部文字化けしたので，encode指定
    lines = file.readlines()

ws.cell(1,1).value = lines[0] #txt file １行目は独立したものだったので．

ttl_list = lines[1].split() #2行目は各列のタイトル．各要素に分割．
for i in range(1,len(lines)): #このfileの行数分繰返し
    for j in range(1,len(ttl_list)+1): #何列のデータか，その分繰返し
        ws.cell(row=i+1, column=j).value = lines[i].split()[j-1] #A2スタートで各リストの要素0から記述

wb.save(xlfile_path)
