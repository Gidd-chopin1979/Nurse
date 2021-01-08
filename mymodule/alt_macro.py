#witのelsx'paste-rowdata'のデータを'macro'に移動
import openpyxl as px
from mymodule import my_round

xxx_name = '002' #ここを範囲指定するかは考え中．その場合は全体をfor
xlfile_path = './Result/1st_Wit_' + xxx_name + '.xlsx'
txt_path = './Data/Wit_001-030/' + xxx_name + 'Data.txt'

wb = px.load_workbook(xlfile_path)
ws = wb['paste-rowdata']
max_row = ws.max_row

wsM = wb['macro']

for i in range(3,max_row+1): #このfileの行数分繰返し
    wsM.cell(row=i, column=2).value = ws.cell(row=i, column=2).value #生データB列をmacroのB列へ
    wsM.cell(row=i, column=4).value = my_round.main(ws.cell(row=i, column=9).value) #生データI列を四捨五入してmacroのD列へ

wb.save(xlfile_path)
