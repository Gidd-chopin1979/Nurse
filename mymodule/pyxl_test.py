import openpyxl as px
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties

#sensorのdataとmovieのdataを取得
nwb = px.Workbook()
nsh = nwb.active
wb1 = px.load_workbook("./data/test1.xlsx")
wb2 = px.load_workbook("./data/test2.xlsx")

#wb1, 2のdataをnwbにコピー
nsh.cell(1,1).value = "data1"
nsh.cell(1,2).value = "data2"
for i in range(2,11):
    nsh.cell(i,1).value = wb1.active.cell(i,1).value
    nsh.cell(i,2).value = wb2.active.cell(i,1).value

#chartを用意
chart = px.chart.ScatterChart(scatterStyle="marker")
x=px.chart.Reference(nsh,min_col=1,max_col=1,min_row=2,max_row=11)
y=px.chart.Reference(nsh,min_col=2,max_col=2,min_row=2,max_row=11)

#--chartのデザイン--#
#グラフのタイトル
chart.title = 'title'

#グラフエリアのサイズ
chart.height = 10
chart.width = 12

#プロットエリアのサイズ・色と枠線
chart.layout = Layout(ManualLayout(x=0,y=0,h=0.8,w=0.8))
chart.plot_area.spPr = GraphicalProperties(noFill = True)
#chart.plot_area.spPr = GraphicalProperties(ln = LineProperties(solidFill = "FFFFFF")

#グリッドラインを消す > どれのことかわかってない
chart.x_axis.majorGridLines = None

#凡例を消す
chart.legend = None

#軸ラベルの名前
chart.x_axis.title = '加速度計の角度'
chart.y_axis.title = '画像の角度'

#最大値・最小値・目盛間隔
chart.x_axis.scaling.min = 0
chart.x_axis.scaling.max = 80
chart.x_axis.majorUnit = 20

chart.y_axis.scaling.min = 0
chart.y_axis.scaling.max = 80
chart.y_axis.majorUnit = 20

#--------------------#

#系列変数を定義
series = px.chart.Series(y,x)

#--ここからデザイン--#
#近似曲線をseriesに追加，描画
series.trendline = px.chart.trendline.Trendline(trendlineType='linear', dispEq=True, dispRSqr=True)

#グラフの線
series.graphicalProperties.line.noFill = True #線をなくす
series.spPr.ln.w = 0.5 * 12700 #線の太さはemuで定義されている．1point=12700 emu

#マーカーを表示，塗潰し無し
series.marker.symbol = "auto"
series.marker.spPr.noFill = True
#--------------------#

#seriesをchartへ渡す
chart.series.append(series)

#グラフの配置:add_chartでchartをB1に置く
nsh.add_chart(chart,"C1")
nwb.save("merge.xlsx")