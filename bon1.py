#witのelsx'paste-rowdata'にwitの生データ.txtを貼り付ける
#読み込み速度の観点から，読み込む列を限定しても良いかもしれない
import csv
import openpyxl as px
from mymodule import hhmmss

xxx_name = ['001','002','003','004','005','006','007','008','009','010','011','012','013','014','015','016','017','018','019','020','021','022','023','024','025','026','027','028','029','030']

for i in range(15,25):
    btemp_path = './Result/xlsx_template/1st_Bonaly_xxx.xlsx'
    bonxl_path = './Result/1st_Bonaly_' + xxx_name[i] + '.xlsx'
    hr_path = './Data/Bonaly_001-030/1st_Bonaly_' + xxx_name[i] + '/' + xxx_name[i] +'.csv'
    acl_path = './Data/Bonaly_001-030/1st_Bonaly_' + xxx_name[i] + '/' + xxx_name[i] +'_acl.csv'

    wb = px.load_workbook(btemp_path)
    wsH = wb['hrate']
    wsA = wb['accel']

    #グラフタイトル名の参照用セルに書き込み
    wb['hrate_graph']['H1'].value = '1次_心拍_' + xxx_name[i]

    #.csvの貼付け
    with open(hr_path, encoding='UTF-8') as file: #txtを開くときと同じ
        line1d = csv.reader(file) #readlineでも可能だが，csv.readerを使う方が勝手がよさそう
        line2d = [row for row in line1d] #二次元配列＝リストのリストとして取得

    for i in range(len(line2d)): #このfileの行数分繰返し
        L_data = line2d[i] 
        for j in range(len(L_data)): #各リストの要素数分だけ繰返し
            wsH.cell(row=i+1, column=j+1).value = line2d[i][j]

    #_acl.csvの貼付け
    with open(acl_path, encoding='UTF-8') as file: #txtを開くときと同じ
        line1D = csv.reader(file) #readlineでも可能だが，csv.readerを使う方が勝手がよさそう
        line2D = [row for row in line1D] #二次元配列＝リストのリストとして取得

    for i in range(len(line2D)): #このfileの行数分繰返し
        L_data = line2D[i] 
        for j in range(len(L_data)): #各リストの要素数分だけ繰返し
            wsA.cell(row=i+1, column=j+1).value = line2D[i][j]

    #'accel' A>G hh:mm:ss
    del line2D[0]
    for i in range(len(line2D)):
        wsA.cell(row=i+2, column=7).value = float(line2D[i][0])/(24*3600)

    #hrateのstart-timeに該当するセルをaccelから探し出し，その経過時刻をhrateのEventに記入，+2してfilldown
    fill_a = px.styles.PatternFill(patternType='solid', fgColor='FFDC00', bgColor='FFDC00') #赤っぽい色に塗りつぶし

    start_list = []

    for i in range(len(line2D)):
        hrate_value = str(hhmmss.main(wsH['A12'].value)) #hrateのhh:mm:ss表記のstart時刻をs表記に変換
        acl_value = str(wsA.cell(row=i+1, column=1).value) #aclのA列，s表記時刻を指定
        if hrate_value in acl_value: #aclの時刻(0.1s刻み)文字列にhrateの時刻(2s刻み)文字列が入っていればtrue
            wsA.cell(row=i+1, column=7).fill = fill_a
            start_list.append(i+1) #list"line2D"は要素を1つ消していて，行数が合わなくなるので+1

    wsA.cell(row=start_list[0], column=9).value = 0 #accelのI列に0を記入
    wsH.cell(row=12, column=2).value = wsA.cell(row=start_list[0], column=8).value #accelのH列「経過時間」をhrateのB12に記入

    for i in range(len(line2d)-12): 
        wsH.cell(row=i+13, column=2).value = wsH.cell(row=i+12, column=2).value + 2 #hrateのB列を12行目から2ずつ足していく

    wb.save(bonxl_path)