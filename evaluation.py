#kinoveaの画像含む，15img/1fileを貼り付ける
#witのelsx'paste-rowdata'にwitの生データ.txtを貼り付ける
#読み込み速度の観点から，読み込む列を限定しても良いかもしれない
import openpyxl as px
from mymodule import my_round

#xxx_name = '002' #ここを範囲指定するかは考え中．その場合は全体をfor
xxx_name = ['001','002','003','004','005','006','007','008','009','010','011','012','013','014','015','016','017','018','019','020','021','022','023','024','025','026','027','028','029','030']

for i in range(0,30):
    print(xxx_name[i])
 
    witxl_path = './Result/1st_Wit_' + xxx_name[i] + '.xlsx'
    eval_path = './Result/Evaluation.xlsx'

    wbW = px.load_workbook(witxl_path)
    wbE = px.load_workbook(eval_path)

    wsW = wbW['corr-factor']
    wsE = wbE['Model']

    G6 = "'[1st_Wit_" + xxx_name[i] + ".xlsx]corr-factor'!$G$6"
    G10 = "'[1st_Wit_" + xxx_name[i] + ".xlsx]corr-factor'!$G$10"
    H6 = "'[1st_Wit_" + xxx_name[i] + ".xlsx]corr-factor'!$H$6"
    H10 = "'[1st_Wit_" + xxx_name[i] + ".xlsx]corr-factor'!$H$10"

    G12 = "'[1st_Wit_" + xxx_name[i] + ".xlsx]corr-factor'!$G$12"
    G16 = "'[1st_Wit_" + xxx_name[i] + ".xlsx]corr-factor'!$G$16"
    H12 = "'[1st_Wit_" + xxx_name[i] + ".xlsx]corr-factor'!$H$12"
    H16 = "'[1st_Wit_" + xxx_name[i] + ".xlsx]corr-factor'!$H$16"    

    #R2 [C列] 1回目/2回目
    wsE.cell(row=i*2+3, column=3).value = "=RSQ(" + H6 + ":" + H10 + ","+ G6 +":"+ G10 + ")"
    wsE.cell(row=i*2+4, column=3).value = "=RSQ(" + H12 + ":" + H16 + ","+ G12 +":"+ G16 + ")"

    #angle D,E列 wit/kinovea
    wsE.cell(row=i*2+3, column=4).value = "=MAX(" + G6+ ":" +G10 +")-MIN("+ G6+ ":" + G10 + ")"
    wsE.cell(row=i*2+4, column=4).value = "=MAX(" + G12+ ":" +G16 +")-MIN("+ G12+ ":" + G16 + ")"

    wsE.cell(row=i*2+3, column=5).value = "=MAX(" + H6+ ":" +H10 +")-MIN("+ H6+ ":" + H10 + ")" #kinovea
    wsE.cell(row=i*2+4, column=5).value = "=MAX(" + H12+ ":" +H16 +")-MIN("+ H12+ ":" + H16 + ")"

    wbE.save(eval_path)

'''
    wsE.cell(row=i*2+3, column=4).value = #"=MIN("+ G6 + ":" + G10 + ")&'－'&MAX(" + G6+ ":" +G10 +")&' ('&(MAX(" + G6+ ":" +G10 +")-MIN("+ G6+ ":" + G10 + "))&')'"
    wsE.cell(row=i*2+4, column=4).value = #"=MIN("+ G12 + ":" + G16 + ")&'－'&MAX(" + G12+ ":" +G16 +")&' ('&(MAX(" + G12+ ":" +G16 +")-MIN("+ G12+ ":" + G16 + "))&')'"

    wsE.cell(row=i*2+3, column=5).value = #"=MIN("+ H6 + ":" + H10 + ")&'－'&MAX(" + H6+ ":" +H10 +")&' ('&(MAX(" + H6+ ":" +H10 +")-MIN("+ H6+ ":" + H10 + "))&')'" #kinovea
    wsE.cell(row=i*2+4, column=5).value = #"=MIN("+ H12 + ":" + H16 + ")&'－'&MAX(" + H12+ ":" +H16 +")&' ('&(MAX(" + H12+ ":" +H16 +")-MIN("+ H12+ ":" + H16 + "))&')'"
'''