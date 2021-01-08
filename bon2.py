#witのelsx'paste-rowdata'にwitの生データ.txtを貼り付ける
#読み込み速度の観点から，読み込む列を限定しても良いかもしれない
import csv
import math
import openpyxl as px
from mymodule import hhmmss
from decimal import Decimal, ROUND_HALF_UP, ROUND_HALF_EVEN #floatの四捨五入

xxx_name = ['001','002','003','004','005','006','007','008','009','010','011','012','013','014','015','016','017','018','019','020','021','022','023','024','025','026','027','028','029','030']

for i in range(15,25):
    print(xxx_name[i])
    witxl_path = './Result/1st_Wit_' + xxx_name[i] + '.xlsx'
    bonxl_path = './Result/1st_Bonaly_' + xxx_name[i] + '.xlsx'
    acl_path = './Data/Bonaly_001-030/1st_Bonaly_' + xxx_name[i] + '/' + xxx_name[i] +'_acl.csv'
    with open(acl_path, encoding='UTF-8') as file: #txtを開くときと同じ
        line1D = csv.reader(file) #readlineでも可能だが，csv.readerを使う方が勝手がよさそう
        line2D = [row for row in line1D] #二次元配列＝リストのリストとして取得

    accel_len = len(line2D)

    wbW = px.load_workbook(witxl_path)
    wsM = wbW['macro']
    maxRowM = wsM.max_row

    wbB = px.load_workbook(bonxl_path)
    wsH = wbB['hrate']
    wsA = wbB['accel']
    ws12 = wbB['1and2']
    wsHG = wbB['hrate_graph']
    maxRowH = wsH.max_row
    maxRowA = wsA.max_row

    fill_a = px.styles.PatternFill(patternType='solid', fgColor='FFDC00', bgColor='FFDC00') #赤っぽい色に塗りつぶし

    se_name = ['1','2','3','4'] #start-endの名前
    se_RowList = [] #区切りを把握しておく
    bw_RowList = [] #witのdataと一致したbonaly accelのデータの行

    #witの区切り番号をse_RowListに格納
    for i in range(1,len(se_name)+1): #1回目開始終了，2回目開始終了の計4つ
        for j in range(1,maxRowM+1):
            if wsM.cell(row=j, column=1).value == i:
                se_RowList.append(j) #区切りに該当する行番号をlistに格納

    #witの区切りデータと一致する時刻をbon accelから探して，区切り番号を割り振る
    for i in range(1,len(se_name)+1):
        wsM_val = float(hhmmss.main(wsM.cell(row=se_RowList[i-1], column=2).value))            
        for j in range(2,accel_len):
            wsA_val = Decimal(wsA.cell(row=j, column=1).value).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP) #roundはfloatを正確に取得できないので，0.5の丸め方が一定方向でない
            if str(wsA_val) in str(wsM_val):
                wsA.cell(row=j, column=7).fill = fill_a 
                bw_RowList.append(j)
                wsA.cell(row=j, column=9).value = i

    #1and2に1,2回目に該当する経過時間と角度のデータを以降
    trial_1 = bw_RowList[1] - bw_RowList[0] + 1 #1回目のデータ数
    trial_2 = bw_RowList[3] - bw_RowList[2] + 1 #2回目のデータ数
    ws12.cell(row=1,column=4).value = str(trial_1)
    ws12.cell(row=1,column=9).value = str(trial_2)

    for i in range(trial_1):
        x = float(wsA.cell(row=i+bw_RowList[0], column=2).value)
        y = float(wsA.cell(row=i+bw_RowList[0], column=4).value)
        ws12.cell(row=i+3, column=3).value = math.atan2(y,x)*180/math.pi*(-1) #腰部角度計算(excel式のコピーを防ぐ)
        ws12.cell(row=i+3, column=2).value = wsA.cell(i+bw_RowList[0], column=8).value #経過時間のコピー

    for j in range(maxRowH-12+1):
        for k in range(trial_1):
            ws12_valA = Decimal(ws12.cell(row=k+3, column=2).value).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP) #float
            wsH_valA = Decimal(wsH.cell(row=j+12, column=2).value).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP) #float
            if ws12_valA == wsH_valA:
                ws12.cell(row=k+3, column=4).value = wsH.cell(row=j+12, column=4).value

    #trial_1と同じ意味だが違う形のfor文
    for i in range(trial_2):
        x = float(wsA.cell(row=i+bw_RowList[2], column=2).value)
        y = float(wsA.cell(row=i+bw_RowList[2], column=4).value)
        ws12.cell(row=i+3, column=8).value = math.atan2(y,x)*180/math.pi*(-1)
        ws12.cell(row=i+3, column=7).value = wsA.cell(i+bw_RowList[3], column=8).value #経過時間のコピー
        for j in range(maxRowH-12+1):
            ws12_valB = Decimal(ws12.cell(row=i+3, column=7).value).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP) #float
            wsH_valB = Decimal(wsH.cell(row=j+12, column=2).value).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP) #float
            if ws12_valB == wsH_valB:
                ws12.cell(row=i+3, column=9).value = wsH.cell(row=j+12, column=4).value

    #1and2 > hrate_graph
    for i in range(trial_1):
        j = str(i+3)
        wsHG.cell(row=i+3,column=2).value = "=VALUE('1and2'!D" + j + ')'
        #wsHG.cell(row=i+3,column=2).number_format = "0" #これをしないとなぜか文字列で入力される
        wsHG.cell(row=i+3,column=3).value = ws12.cell(row=i+3,column=5).value
        

    for i in range(trial_2):
        j = str(i+3)
        wsHG.cell(row=i+3,column=5).value = "=VALUE('1and2'!I" + j + ')'
        wsHG.cell(row=i+3,column=6).value = ws12.cell(row=i+3,column=10).value

    wbB.save(bonxl_path)